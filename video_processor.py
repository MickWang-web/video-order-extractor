"""
视频处理模块：拆帧 + 视觉识别 + 数据汇总
"""
import os
import re
import json
import subprocess
import base64
from pathlib import Path
from typing import Optional

from openai import OpenAI
from config import VISION_API_BASE, VISION_API_KEY, VISION_MODEL, FRAMES_DIR, OUTPUT_DIR


# ========== 1. 视频拆帧 ==========

def extract_frames(video_path: str, fps: float = 0.5, output_dir: Optional[str] = None) -> list[str]:
    """
    使用 ffmpeg 从视频中提取关键帧
    
    Args:
        video_path: 视频文件路径
        fps: 每秒提取帧数，0.5表示每2秒1帧
        output_dir: 帧图片输出目录
    
    Returns:
        帧图片路径列表
    """
    if output_dir is None:
        output_dir = FRAMES_DIR
    
    os.makedirs(output_dir, exist_ok=True)
    
    output_pattern = os.path.join(output_dir, "frame_%04d.jpg")
    
    cmd = [
        "ffmpeg",
        "-y",           # 覆盖已有文件
        "-i", video_path,
        "-vf", f"fps={fps}",
        "-q:v", "2",    # JPEG质量(2最好，31最差)
        output_pattern
    ]
    
    result = subprocess.run(cmd, capture_output=True, text=True, timeout=120)
    
    if result.returncode != 0:
        raise RuntimeError(f"ffmpeg执行失败: {result.stderr}")
    
    # 收集生成的帧文件
    frames = sorted([
        os.path.join(output_dir, f) 
        for f in os.listdir(output_dir) 
        if f.startswith("frame_") and f.endswith(".jpg")
    ])
    
    return frames


def smart_extract_frames(video_path: str, output_dir: Optional[str] = None) -> list[str]:
    """
    智能拆帧：根据视频时长自动选择合适的帧率
    - ≤10秒：每秒1帧
    - 10-60秒：每2秒1帧
    - 60-300秒：每3秒1帧
    - >300秒：每5秒1帧
    """
    # 获取视频时长
    cmd = [
        "ffprobe", "-v", "quiet",
        "-print_format", "json",
        "-show_format", video_path
    ]
    result = subprocess.run(cmd, capture_output=True, text=True)
    duration = float(json.loads(result.stdout)["format"]["duration"])
    
    if duration <= 10:
        fps = 1.0
    elif duration <= 60:
        fps = 0.5
    elif duration <= 300:
        fps = 1.0 / 3
    else:
        fps = 0.2
    
    print(f"视频时长: {duration:.1f}秒, 拆帧频率: {fps}fps")
    return extract_frames(video_path, fps, output_dir)


# ========== 2. 视觉大模型识别 ==========

def get_vision_client() -> OpenAI:
    """获取视觉大模型客户端（OpenAI兼容接口）"""
    return OpenAI(
        base_url=VISION_API_BASE,
        api_key=VISION_API_KEY
    )


def image_to_base64(image_path: str) -> str:
    """将图片转为base64编码"""
    with open(image_path, "rb") as f:
        return base64.b64encode(f.read()).decode("utf-8")


# 识别单张图片的提示词
OCR_PROMPT = """你是一个专业的订单信息提取助手。请仔细观察这张手机截图，这是购物App的"我的订单"页面。

请提取页面中每一条订单的以下信息，以JSON数组格式返回：
- order_id: 订单编号（完整数字串）
- status: 订单状态（如"已完成"、"已退款"等）
- datetime: 下单时间（格式 YYYY/MM/DD HH:MM:SS）
- amount: 金额（数字，退款为负数，如 -145.90）
- tags: 标签（如"亲友卡"等，没有则为空字符串）

注意事项：
1. 只返回页面上能清晰看到的订单，看不清或截断的不完整订单也要提取，尽量补全
2. 金额要去掉¥符号，只保留数字
3. 退款订单金额为负数
4. 订单编号尽量提取完整
5. 只返回JSON，不要其他文字

返回格式示例：
```json
[
  {"order_id": "557100000000060717202605242120", "status": "已完成", "datetime": "2026/05/24 21:19:36", "amount": 52.80, "tags": ""},
  {"order_id": "557100000008300049202605112104", "status": "已退款", "datetime": "2026/05/11 21:06:17", "amount": -145.90, "tags": ""}
]
```
"""


def recognize_frame(client: OpenAI, image_path: str) -> list[dict]:
    """
    调用视觉大模型识别单帧图片中的订单信息
    
    Args:
        client: OpenAI客户端
        image_path: 帧图片路径
    
    Returns:
        识别到的订单列表
    """
    b64 = image_to_base64(image_path)
    
    try:
        response = client.chat.completions.create(
            model=VISION_MODEL,
            messages=[
                {
                    "role": "user",
                    "content": [
                        {"type": "text", "text": OCR_PROMPT},
                        {
                            "type": "image_url",
                            "image_url": {
                                "url": f"data:image/jpeg;base64,{b64}"
                            }
                        }
                    ]
                }
            ],
            max_tokens=2000,
            temperature=0.1  # 低温度保证提取准确性
        )
        
        text = response.choices[0].message.content
        
        # 从返回文本中提取JSON
        orders = parse_orders_from_text(text)
        return orders
        
    except Exception as e:
        import traceback
        error_msg = f"识别帧 {image_path} 失败: {type(e).__name__}: {e}"
        print(error_msg)
        print("完整错误堆栈:")
        print(traceback.format_exc())
        raise


def parse_orders_from_text(text: str) -> list[dict]:
    """
    从模型返回文本中解析JSON订单数据
    模型返回可能包含 ```json ... ``` 包裹，需要提取
    """
    # 尝试提取 ```json ... ``` 中的内容
    json_match = re.search(r'```(?:json)?\s*([\s\S]*?)```', text)
    if json_match:
        json_str = json_match.group(1).strip()
    else:
        json_str = text.strip()
    
    try:
        orders = json.loads(json_str)
        if isinstance(orders, list):
            return orders
    except json.JSONDecodeError:
        pass
    
    # 如果整体解析失败，尝试找 [ ... ] 的部分
    bracket_match = re.search(r'\[[\s\S]*\]', text)
    if bracket_match:
        try:
            orders = json.loads(bracket_match.group())
            if isinstance(orders, list):
                return orders
        except json.JSONDecodeError:
            pass
    
    print(f"无法解析模型返回: {text[:200]}")
    return []


def recognize_all_frames(frame_paths: list[str], batch_size: int = 3) -> list[dict]:
    """
    批量识别所有帧
    
    Args:
        frame_paths: 帧图片路径列表
        batch_size: 并发识别的批次大小（避免API限流）
    
    Returns:
        所有识别到的订单（含重复）
    """
    import time
    
    client = get_vision_client()
    all_orders = []
    
    total = len(frame_paths)
    for i, frame_path in enumerate(frame_paths):
        print(f"正在识别第 {i+1}/{total} 帧: {os.path.basename(frame_path)}")
        
        orders = recognize_frame(client, frame_path)
        all_orders.extend(orders)
        print(f"  识别到 {len(orders)} 条订单")
        
        # 简单限流：每批之间等待
        if (i + 1) % batch_size == 0 and i + 1 < total:
            print("  等待1秒避免限流...")
            time.sleep(1)
    
    return all_orders


# ========== 3. 数据去重与汇总 ==========

def deduplicate_orders(orders: list[dict]) -> list[dict]:
    """
    去重：同一订单号可能出现多次（在多帧中），只保留信息最完整的那条
    """
    seen = {}
    
    for o in orders:
        oid = str(o.get("order_id", "")).replace(" ", "").strip()
        if not oid or len(oid) < 10:
            continue
        
        # 如果该订单号已存在，保留字段更完整的那个
        if oid in seen:
            existing = seen[oid]
            # 计算非空字段数
            existing_score = sum(1 for v in existing.values() if v is not None and v != "" and v != 0)
            new_score = sum(1 for k, v in o.items() if v is not None and v != "" and v != 0)
            if new_score > existing_score:
                seen[oid] = o
        else:
            seen[oid] = o
    
    # 按时间倒序排列
    deduped = list(seen.values())
    deduped.sort(key=lambda x: x.get("datetime", ""), reverse=True)
    
    return deduped


def generate_summary(orders: list[dict]) -> dict:
    """
    生成汇总统计
    """
    completed = [o for o in orders if o.get("status") == "已完成"]
    refunded = [o for o in orders if o.get("status") == "已退款"]
    qyk_orders = [o for o in orders if "亲友卡" in (o.get("tags", ""))]
    
    completed_amount = sum(float(o.get("amount", 0)) for o in completed)
    refund_amount = sum(float(o.get("amount", 0)) for o in refunded)
    qyk_amount = sum(float(o.get("amount", 0)) for o in qyk_orders if o.get("status") == "已完成")
    
    # 月度汇总
    from collections import defaultdict
    monthly = defaultdict(lambda: {"count": 0, "completed": 0, "refund": 0, 
                                    "completed_amount": 0.0, "refund_amount": 0.0})
    for o in orders:
        month = o.get("datetime", "")[:7]  # YYYY/MM
        if not month or len(month) < 7:
            continue
        monthly[month]["count"] += 1
        if o.get("status") == "已完成":
            monthly[month]["completed"] += 1
            monthly[month]["completed_amount"] += float(o.get("amount", 0))
        elif o.get("status") == "已退款":
            monthly[month]["refund"] += 1
            monthly[month]["refund_amount"] += float(o.get("amount", 0))
    
    monthly_summary = {}
    for month, data in sorted(monthly.items(), reverse=True):
        monthly_summary[month] = {
            "count": data["count"],
            "completed": data["completed"],
            "refund": data["refund"],
            "completed_amount": round(data["completed_amount"], 2),
            "refund_amount": round(data["refund_amount"], 2),
            "net_amount": round(data["completed_amount"] + data["refund_amount"], 2)
        }
    
    return {
        "total_orders": len(orders),
        "completed_count": len(completed),
        "refund_count": len(refunded),
        "completed_amount": round(completed_amount, 2),
        "refund_amount": round(refund_amount, 2),
        "net_amount": round(completed_amount + refund_amount, 2),
        "qyk_count": len(qyk_orders),
        "qyk_completed_amount": round(qyk_amount, 2),
        "monthly_summary": monthly_summary
    }


# ========== 4. 生成Excel ==========

def generate_excel(orders: list[dict], summary: dict, output_filename: str) -> str:
    """
    生成Excel报表
    
    Returns:
        Excel文件路径
    """
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    import traceback
    
    try:
        wb = openpyxl.Workbook()
        
        # ---- Sheet1: 订单明细 ----
        ws = wb.active
        ws.title = "订单明细"
        
        headers = ["序号", "订单号", "订单状态", "下单时间", "金额(元)", "标签"]
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        )
        
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        for row_idx, order in enumerate(orders, 2):
            values = [
                row_idx - 1,
                str(order.get("order_id", "")),
                order.get("status", ""),
                order.get("datetime", ""),
                float(order.get("amount", 0)),
                order.get("tags", "")
            ]
            for col_idx, val in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.border = thin_border
                if col_idx == 5:
                    cell.number_format = '#,##0.00'
                if col_idx == 3 and val == "已退款":
                    cell.font = Font(color="FF0000")
                if col_idx == 5 and isinstance(val, (int, float)) and val < 0:
                    cell.font = Font(color="FF0000")
        
        ws.column_dimensions['A'].width = 6
        ws.column_dimensions['B'].width = 38
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 22
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 10
        
        # ---- Sheet2: 月度汇总 ----
        ws2 = wb.create_sheet("月度汇总")
        month_headers = ["月份", "订单数", "已完成", "已退款", "已完成金额", "退款金额", "净消费"]
        for col, h in enumerate(month_headers, 1):
            cell = ws2.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
        
        for i, (month, data) in enumerate(summary.get("monthly_summary", {}).items(), 2):
            ws2.cell(row=i, column=1, value=month)
            ws2.cell(row=i, column=2, value=data["count"])
            ws2.cell(row=i, column=3, value=data["completed"])
            ws2.cell(row=i, column=4, value=data["refund"])
            ws2.cell(row=i, column=5, value=data["completed_amount"])
            ws2.cell(row=i, column=6, value=data["refund_amount"])
            ws2.cell(row=i, column=7, value=data["net_amount"])
            for col in range(5, 8):
                ws2.cell(row=i, column=col).number_format = '#,##0.00'
        
        for col in range(1, 8):
            ws2.column_dimensions[chr(64 + col)].width = 14
        
        # ---- Sheet3: 汇总统计 ----
        ws3 = wb.create_sheet("汇总统计")
        stat_items = [
            ("总订单数", summary["total_orders"]),
            ("已完成", summary["completed_count"]),
            ("已退款", summary["refund_count"]),
            ("已完成金额合计", summary["completed_amount"]),
            ("退款金额合计", summary["refund_amount"]),
            ("实际净消费", summary["net_amount"]),
            ("亲友卡订单数", summary["qyk_count"]),
            ("亲友卡已完成金额", summary["qyk_completed_amount"]),
        ]
        for i, (label, val) in enumerate(stat_items, 1):
            ws3.cell(row=i, column=1, value=label).font = Font(bold=True)
            ws3.cell(row=i, column=2, value=val)
        ws3.column_dimensions['A'].width = 18
        ws3.column_dimensions['B'].width = 16
        
        # 保存
        output_path = os.path.join(OUTPUT_DIR, output_filename)
        wb.save(output_path)
        return output_path
    except Exception as e:
        error_msg = f"生成Excel失败: {type(e).__name__}: {e}"
        print(error_msg)
        print("完整错误堆栈:")
        print(traceback.format_exc())
        raise


# ========== 5. 完整处理流程 ==========

def process_video(video_path: str, fps: Optional[float] = None) -> dict:
    """
    完整的视频订单提取流程
    
    Args:
        video_path: 视频文件路径
        fps: 拆帧频率，None为自动
    
    Returns:
        处理结果字典
    """
    from datetime import datetime
    
    print("=" * 50)
    print("开始处理视频...")
    
    # Step 1: 拆帧
    print("\n[1/4] 提取视频帧...")
    if fps:
        frames = extract_frames(video_path, fps)
    else:
        frames = smart_extract_frames(video_path)
    print(f"共提取 {len(frames)} 帧")
    
    if not frames:
        return {"success": False, "error": "未能提取到任何帧，请检查视频文件"}
    
    # Step 2: 逐帧识别
    print("\n[2/4] 识别订单信息...")
    raw_orders = recognize_all_frames(frames)
    print(f"原始识别到 {len(raw_orders)} 条订单（含重复）")
    
    # Step 3: 去重汇总
    print("\n[3/4] 去重汇总...")
    orders = deduplicate_orders(raw_orders)
    summary = generate_summary(orders)
    print(f"去重后 {summary['total_orders']} 条订单")
    print(f"已完成 {summary['completed_count']} 笔，退款 {summary['refund_count']} 笔")
    print(f"净消费: ¥{summary['net_amount']:.2f}")
    
    # Step 4: 生成Excel
    print("\n[4/4] 生成Excel...")
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    excel_filename = f"orders_{timestamp}.xlsx"
    excel_path = generate_excel(orders, summary, excel_filename)
    print(f"Excel已保存: {excel_path}")
    
    # 清理临时帧文件
    for f in frames:
        try:
            os.remove(f)
        except:
            pass
    
    print("\n处理完成！")
    print("=" * 50)
    
    return {
        "success": True,
        "data": {
            **summary,
            "orders": orders,
            "excel_filename": excel_filename
        }
    }
