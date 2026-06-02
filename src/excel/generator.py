"""
Excel报表生成模块
"""

import os
import traceback
from typing import Dict, List, Any

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

from src.core.config import OUTPUT_DIR


def generate_excel(orders: List[Dict[str, Any]], summary: Dict[str, Any], output_filename: str) -> str:
    """
    生成Excel报表

    Args:
        orders: 订单列表
        summary: 汇总统计
        output_filename: 输出文件名

    Returns:
        Excel文件路径
    """
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = "订单明细"

    headers = ["序号", "订单号", "订单状态", "下单时间", "金额(元)", "标签"]
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    for row_idx, order in enumerate(orders, 2):
        values = [
            row_idx - 1,
            str(order.get("order_id", "")),
            order.get("status", ""),
            order.get("datetime", ""),
            float(order.get("amount", 0)),
            order.get("tags", "")
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            if col_idx == 5:
                cell.number_format = '#,##0.00'
            if col_idx == 3 and val == "已退款":
                cell.font = Font(color="FF0000")
            if col_idx == 5 and isinstance(val, (int, float)) and val < 0:
                cell.font = Font(color="FF0000")

    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 38
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 22
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 10

    ws2 = wb.create_sheet("月度汇总")
    month_headers = ["月份", "订单数", "已完成", "已退款", "已完成金额", "退款金额", "净消费"]
    for col, h in enumerate(month_headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill

    for i, (month, data) in enumerate(summary.get("monthly_summary", {}).items(), 2):
        ws2.cell(row=i, column=1, value=month)
        ws2.cell(row=i, column=2, value=data["count"])
        ws2.cell(row=i, column=3, value=data["completed"])
        ws2.cell(row=i, column=4, value=data["refund"])
        ws2.cell(row=i, column=5, value=data["completed_amount"])
        ws2.cell(row=i, column=6, value=data["refund_amount"])
        ws2.cell(row=i, column=7, value=data["net_amount"])
        for col in range(5, 8):
            ws2.cell(row=i, column=col).number_format = '#,##0.00'

    for col in range(1, 8):
        ws2.column_dimensions[chr(64 + col)].width = 14

    ws3 = wb.create_sheet("汇总统计")
    stat_items = [
        ("总订单数", summary["total_orders"]),
        ("已完成", summary["completed_count"]),
        ("已退款", summary["refund_count"]),
        ("已完成金额合计", summary["completed_amount"]),
        ("退款金额合计", summary["refund_amount"]),
        ("实际净消费", summary["net_amount"]),
        ("亲友卡订单数", summary["qyk_count"]),
        ("亲友卡已完成金额", summary["qyk_completed_amount"]),
    ]
    for i, (label, val) in enumerate(stat_items, 1):
        ws3.cell(row=i, column=1, value=label).font = Font(bold=True)
        ws3.cell(row=i, column=2, value=val)
    ws3.column_dimensions['A'].width = 18
    ws3.column_dimensions['B'].width = 16

    output_path = os.path.join(OUTPUT_DIR, output_filename)
    wb.save(output_path)
    return output_path
